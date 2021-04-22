import re
from copy import copy as cp
from pathlib import Path

import openpyxl as xl
from datascroller import scroll
from openpyxl.formatting.formatting import ConditionalFormattingList
from openpyxl.formula.translate import Translator
from openpyxl.worksheet.cell_range import CellRange, MultiCellRange
from openpyxl.worksheet.table import Table, TableStyleInfo

from .__init__ import *
from .. import styles as st

log = getlog(__name__)
p = Path.home() / 'desktop/Activity Record - Jayme Gordon.xlsx'


class ExcelModel():
    def __init__(self):
        wb = xl.load_workbook(p)
        ws = wb['ActLog'] # default table
        tbl = ws.tables['Table1']
        max_row = ws.max_row
        new_row = ws.max_row + 1

        m_formula_cols = dict(actlog=2)
        
        f.set_self(vars())
    
    def get_vars(self):
        return self.p, self.wb, self.ws, self.tbl
    
    def get_tbl(self, ws):
        """Return first table of worksheet"""
        return ws.tables[list(ws.tables)[0]]
    
    def close(self, display=True, **kw):
        if display:
            self.display_df(**kw)

        self.wb.save(p)
        self.wb.close()
    
    def add_row(self, m: dict, name: str='ActLog'):
        """Add row(s) to excel table
        
        Parameters
        ----------
        m : dict | list
        """
        ws = self.wb[name]
        tbl = self.get_tbl(ws=ws)
        cols = f.lower_cols(tbl.column_names)
        row = ws.max_row

        # convert dict m to list
        lst = m if isinstance(m, list) else [m]
        n = len(lst)

        # adjust table range and cond formats
        tbl.ref = inc_rng(tbl.ref, n=n)
        max_col = len(tbl.column_names)
        self.adjust_cond_formats(n=n, name=name)
        
        # fill formula cols down
        formula_col = self.m_formula_cols.get(name.lower(), None)
        if not formula_col is None:
            self.fill_formula(col=formula_col, n=n)

        for m in lst:
            row += 1

            for col_name, val in m.items():
                col = cols.index(col_name) + 1 # get col int from colname
                ws.cell(row, col).value = val

            for col in range(max_col):
                col += 1
                cell = ws.cell(row - 1, col)
                new_cell = ws.cell(row, col)

                new_cell.alignment = cp(cell.alignment)
                new_cell.number_format = cp(cell.number_format)
    
    def delete_row(self, n: int=1):
        ws, tbl, row = self.ws, self.tbl, self.max_row

        for _ in range(n):
            ws.delete_rows(idx=ws.max_row)
        
        tbl.ref = inc_rng(tbl.ref, n=-1 * n)
        self.adjust_cond_formats(n=-1 * n)
    
    def fill_formula(self, col, n: int=1, name: str='ActLog'):
        """Copy cell formula down in col"""
        ws = self.wb[name]
        # NOTE ws max row kinda messy, checks max row with table/formats, NOT values
        start_row = self.max_row

        for i in range(n):
            c = ws.cell(start_row + i, col)
            ws.cell(start_row + i + 1, col).value = Translator(c.value, origin=c.coordinate) \
                .translate_formula(inc_rng(c.coordinate))

    def get_df(self, name='ActLog'):
        """Return df from excel table"""
        ws = self.wb[name]
        tbl = self.get_tbl(ws=ws)
        max_col = len(tbl.column_names)

        data = np.array(list(ws.values)[1:])[:, :max_col]
        cols = next(ws.values)[:max_col]

        drop_cols = dict(actlog=['sum_duration']) \
            .get(name.lower(), None)

        df = pd.DataFrame(data=data, columns=cols) \
            .pipe(f.lower_cols) \
        
        if 'date' in df.columns:
            df = df.sort_values('date')
        
        if not drop_cols is None:
            df = df.drop(columns=drop_cols)

        if name == 'ActLog':
            # add sum duration to first row per date
            df_sum = df[['date']].drop_duplicates('date', keep='first') \
                .reset_index() \
                .assign(day=lambda x: x.date.dt.strftime('%a')) \
                .merge(right=df.groupby('date', as_index=False)[['duration']].sum()) \
                .set_index('index')[['duration', 'day']] \
                .rename(columns=dict(duration='sum'))

            cols = ['date', 'day', 'sum', 'duration', 'task', 'task_type']
            df = df.merge(right=df_sum, how='left', left_index=True, right_index=True) \
                [cols]

        return df
    
    def get_df_sum(self, n: int=10):
        """Return df of summary durations for last n dates"""

        df_sum = self.get_df(name='ActLog') \
            .groupby('date', as_index=False)[['date', 'day', 'sum']].first() \
            .set_index('date')

        d = dt.now().date()
        rng = pd.date_range(d + delta(days=-n), d)

        return pd.DataFrame(index=rng) \
            .rename_axis('date') \
            .merge(right=df_sum, left_index=True, right_index=True, how='left') \
            .reset_index() \
            .assign(
                day=lambda x: x.date.dt.strftime('%a'),
                sum=lambda x: x['sum'].fillna(0).astype(float))
    
    def highlight_sum(self, df, color=160):
        """Highlight sum red if less than 8 hrs"""

        # filter out weekends
        mask = (~df.date.dt.strftime('%a').isin(('Sat', 'Sun'))) & (df['sum'] < 8)
        
        data = np.vstack(
            [np.array([''] * df.shape[0]),
            np.where(mask, f'color: {color}', '')]).T
        
        return pd.DataFrame(columns=df.columns, index=df.index, data=data)
    
    def show_df_sum(self, n: int=10):
        df = self.get_df_sum(n=n) \
            .tail(n)
        
        style = df.style \
            .apply(self.highlight_sum, subset=['date', 'sum'], axis=None) \
            .applymap(lambda x: 'color: 226' if x in ('Sat', 'Sun') else '', subset='day') \
            # .applymap(lambda x: 'color: royalblue', subset='date')

        self.display_df(df=df, n=n, style=style, showindex=False, floatfmt='.1f')
    
    def show_like(self, s):
        """Show last n items containing search string in task"""
        df = self.get_df()
        df = df[df.task.str.contains(s, case=False, na=False)]
        scroll(df)

    def display_df(self, n: int=None, df=None, style=None, d_lower=None, **kw):
        """Display last n rows of table"""
        if n is None:
            n = 10

        if df is None:
            df = self.get_df()

            # show last n rows, or whichever date added + n
            if not d_lower is None:
                df2 = df[df.date >= d_lower] \
                    .head(n)
                
                df = df2 if df2.shape[0] == n else df.tail(n)
            else:
                df = df.tail(n)

            style = df.style \
                .apply(self.highlight_sum, subset=['date', 'sum'], axis=None) \
                .apply(st.highlight_alternating, axis=None, color='royalblue', anchor_col='date', is_hex=False)
        
        if not style is None:
            df = st.terminal_color(df, style)

        f.terminal_df(df, **kw)

    def adjust_cond_formats(self, n: int=1, name: str='ActLog'):
        """Extend conditional formatting by one row
        - have to delete then re-add cond formats"""
        ws = self.wb[name]
        orig_format = cp(ws.conditional_formatting)
        ws.conditional_formatting = ConditionalFormattingList() # clear formats

        for cond in orig_format:
            ws.conditional_formatting.add(
                inc_rng(cond.sqref, n=n),
                cond.cfRule[0])

def get_matching_task(df, task):
    """Return category if task exists already"""
    df = df[df.task.str.lower() == task]

    try:
        return df.iloc[0, df.columns.get_loc('task_type')].title()
    except:
        return ''

def add_defaults(d: dt=None, em=None):
    """Add default rows for all dates till current date with 0 time"""
    # get df
    # groupby date
    # get date_range btwn now and start
    # merge grouped data
    # filter dates = 0 or Nan

    if em is None:
        em = ExcelModel()
    
    if not d is None:
        # init specific day
        day = d.strftime('%a')

        df = em.get_df(name='Default') \
            .pipe(lambda df: df[df.day==day]) \
            .drop(columns='day') \
            .assign(
                duration=lambda x: x.duration.astype(float),
                date=d.date())
    
    else:
        # init all blank days
        d = dt.now().date()
        n = 90
        d_lower = d + delta(days=-n)
        rng = pd.date_range(d_lower, d)

        df_default = em.get_df(name='Default') # defaults per day

        # get sum duration from last n days per day
        df_sum = em.get_df('ActLog') \
            .pipe(lambda df: df[df.date.dt.date >= d_lower]) \
            .groupby('date', as_index=False)[['duration']].sum() \
            .rename(columns=dict(duration='sum'))

        # merge default values for all days with 0 duration
        df = pd.DataFrame(dict(date=rng)) \
            .assign(day=lambda x: x.date.dt.strftime('%a')) \
            .merge(right=df_sum, on='date', how='left') \
            .fillna(0) \
            .merge(right=df_default, on='day', how='outer') \
            .dropna() \
            .pipe(lambda df: df[df['sum'] == 0]) \
            .drop(columns=['day', 'sum']) \
            .sort_values('date') \
            .assign(
                duration=lambda x: x.duration.astype(float))

    lst = list(df.to_dict(orient='index').values())
    em.add_row(m=lst)
    em.close()

    n_dates = df.groupby('date').size().shape[0]
    print('\n')
    log.info(f'Dates initialized: {n_dates}')

def update_time(task, duration: float=1.0, d: dt=None, category=None, em=None, n: int=None):
    """Add row to table with new task data"""
    if em is None:
        em = ExcelModel()
    
    if d is None:
        d = dt.now().date()

    if isinstance(d, str):
        d = dt.strptime(d, '%Y-%m-%d')
    
    if category is None:
        category = get_matching_task(df=em.get_df(), task=task)

    m = dict(
        date=d,
        duration=duration,
        task=f'{task[0].upper()}{task[1:]}',
        task_type=category.title())

    em.add_row(m=m, name='ActLog')
    em.close(n=n, d_lower=f.convert_date(d))

def delete_row(n: int=1, em=None):
    if em is None:
        em = ExcelModel()

    em.delete_row(n=n)
    em.close()

def inc_rng(rng: str, n: int=1):
    """increment or expand cell range by 1 row"""
    expr = r'\d+$'
    rng = str(rng)
    row = int(re.search(expr, rng)[0])
    return re.sub(expr, str(row + n), rng)
